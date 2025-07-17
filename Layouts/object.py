import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import math
import copy

def parse_svg_file(filename):
    """Parse SVG file and extract all elements"""
    try:
        tree = ET.parse(filename)
        root = tree.getroot()
        return root
    except Exception as e:
        print(f"Error parsing SVG file: {e}")
        return None

def extract_all_transforms(transform_str):
    """Extract all transformation values from a transform string"""
    transforms = {
        'translate': (0.0, 0.0),
        'scale': (1.0, 1.0),
        'rotate': 0.0,
        'matrix': None
    }
    
    if not transform_str:
        return transforms
    
    # Translate
    translate_match = re.search(r'translate\s*\(\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)\s*\)', transform_str)
    if translate_match:
        transforms['translate'] = (float(translate_match.group(1)), float(translate_match.group(2)))
    
    # Scale
    scale_match = re.search(r'scale\s*\(\s*([+-]?\d*\.?\d+)(?:\s*[,\s]\s*([+-]?\d*\.?\d+))?\s*\)', transform_str)
    if scale_match:
        sx = float(scale_match.group(1))
        sy = float(scale_match.group(2)) if scale_match.group(2) else sx
        transforms['scale'] = (sx, sy)
    
    # Matrix
    matrix_match = re.search(r'matrix\s*\(\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)\s*\)', transform_str)
    if matrix_match:
        transforms['matrix'] = [float(matrix_match.group(i)) for i in range(1, 7)]
    
    return transforms

def apply_transform_to_point(x, y, transforms):
    """Apply transformation to a point"""
    # Apply scale
    sx, sy = transforms['scale']
    x *= sx
    y *= sy
    
    # Apply matrix if present
    if transforms['matrix']:
        a, b, c, d, e, f = transforms['matrix']
        new_x = a * x + c * y + e
        new_y = b * x + d * y + f
        x, y = new_x, new_y
    else:
        # Apply translate
        tx, ty = transforms['translate']
        x += tx
        y += ty
    
    return x, y

def get_path_bounds(path_data):
    """Get approximate bounds of a path by parsing move commands"""
    if not path_data:
        return 0, 0
    
    # Extract all M (move) commands to find starting positions
    moves = re.findall(r'[Mm]\s*([+-]?\d*\.?\d+)\s*[,\s]\s*([+-]?\d*\.?\d+)', path_data)
    
    if moves:
        # Use the first move command as the position
        x = float(moves[0][0])
        y = float(moves[0][1])
        return x, y
    
    return 0, 0

def get_element_position(element, root):
    """Get the absolute position of an element including its own coordinates"""
    # Build parent map
    parent_map = {c: p for p in root.iter() for c in p}
    
    # Get path data if this is a path element or has path children
    local_x, local_y = 0, 0
    
    # If element has path children, get position from path data
    for child in element:
        if child.tag.endswith('path'):
            path_data = child.get('d', '')
            if path_data:
                local_x, local_y = get_path_bounds(path_data)
                break
    
    # If element itself is a path
    if element.tag.endswith('path'):
        path_data = element.get('d', '')
        if path_data:
            local_x, local_y = get_path_bounds(path_data)
    
    # Traverse up the tree and accumulate transforms
    current = element
    x_total, y_total = local_x, local_y
    
    while current is not None and current != root:
        transform_str = current.get('transform', '')
        if transform_str:
            transforms = extract_all_transforms(transform_str)
            x_total, y_total = apply_transform_to_point(x_total, y_total, transforms)
        
        # Move to parent
        current = parent_map.get(current)
    
    return x_total, y_total

def get_element_bounds(element):
    """Get the bounding box of an element by analyzing all its coordinates"""
    min_x, min_y = float('inf'), float('inf')
    max_x, max_y = float('-inf'), float('-inf')
    
    # Process all child elements
    for child in element.iter():
        # Handle line elements
        if child.tag.endswith('line'):
            x1 = float(child.get('x1', 0))
            y1 = float(child.get('y1', 0))
            x2 = float(child.get('x2', 0))
            y2 = float(child.get('y2', 0))
            
            min_x = min(min_x, x1, x2)
            max_x = max(max_x, x1, x2)
            min_y = min(min_y, y1, y2)
            max_y = max(max_y, y1, y2)
        
        # Handle rect elements
        elif child.tag.endswith('rect'):
            x = float(child.get('x', 0))
            y = float(child.get('y', 0))
            width = float(child.get('width', 0))
            height = float(child.get('height', 0))
            
            min_x = min(min_x, x)
            max_x = max(max_x, x + width)
            min_y = min(min_y, y)
            max_y = max(max_y, y + height)
        
        # Handle circle elements
        elif child.tag.endswith('circle'):
            cx = float(child.get('cx', 0))
            cy = float(child.get('cy', 0))
            r = float(child.get('r', 0))
            
            min_x = min(min_x, cx - r)
            max_x = max(max_x, cx + r)
            min_y = min(min_y, cy - r)
            max_y = max(max_y, cy + r)
        
        # Handle path elements
        elif child.tag.endswith('path'):
            path_data = child.get('d', '')
            if path_data:
                # Extract all numeric values from path
                numbers = re.findall(r'[+-]?\d*\.?\d+', path_data)
                if len(numbers) >= 2:
                    # Process pairs of coordinates
                    for i in range(0, len(numbers)-1, 2):
                        x = float(numbers[i])
                        y = float(numbers[i+1])
                        min_x = min(min_x, x)
                        max_x = max(max_x, x)
                        min_y = min(min_y, y)
                        max_y = max(max_y, y)
    
    # If no bounds found, return default size
    if min_x == float('inf'):
        return 0, 0
    
    width = max_x - min_x
    height = max_y - min_y
    
    return width, height

def is_row(element):
    """Check if an element is a row based on its characteristics"""
    # Check if it's a group element
    if not element.tag.endswith('g'):
        return False
    
    # Check fill color - rows are beige/cream colored with opacity
    g_fill = element.get('fill', '')
    fill_opacity = element.get('fill-opacity', '1')
    
    # Check for the specific beige/cream color (rgb(224,224,215))
    is_beige = 'rgb(224,224,215)' in g_fill
    
    # Check for semi-transparent fill
    try:
        opacity = float(fill_opacity)
        is_semi_transparent = 0.3 <= opacity <= 0.5
    except:
        is_semi_transparent = False
    
    if not (is_beige and is_semi_transparent):
        return False
    
    # Check for path children that form rectangles
    for child in element:
        if child.tag.endswith('path'):
            path_data = child.get('d', '')
            # Row paths typically contain Z (close path) and form rectangles
            # They should have L (line) commands and Z to close
            if 'Z' in path_data and 'L' in path_data:
                # Check if it's a vertical rectangle (typical for rows)
                numbers = re.findall(r'\d+', path_data)
                if len(numbers) >= 4:
                    # Extract coordinates to check if it's vertical
                    coords = [int(n) for n in numbers[:8] if n.isdigit()]
                    if len(coords) >= 4:
                        # Check if height > width (vertical rectangle)
                        width = abs(coords[4] - coords[0]) if len(coords) > 4 else 0
                        height = abs(coords[3] - coords[1]) if len(coords) > 3 else 0
                        if height > width * 2:  # Vertical rectangle
                            return True
    
    return False

def find_rows(root):
    """Find all rows in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    rows = []
    processed_groups = set()
    
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        if is_row(g):
            processed_groups.add(g_id)
            
            # Get position
            x, y = get_element_position(g, root)
            
            # Get dimensions
            width, height = get_element_bounds(g)
            
            rows.append({
                'element': g,
                'x': x,
                'y': y,
                'width': width,
                'height': height
            })
            
            # Mark child groups as processed
            for child_g in g.findall('.//svg:g', namespaces):
                processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_rows = []
    positions_seen = set()
    
    for row in rows:
        pos_key = f"{row['x']:.1f},{row['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_rows.append(row)
    
    return unique_rows

def is_cooling_tile_pattern(path_data):
    """Check if a path represents a cooling tile pattern"""
    if not path_data:
        return False
    
    upper_data = path_data.upper()
    c_count = upper_data.count('C')
    m_count = upper_data.count('M')
    z_count = upper_data.count('Z')
    
    # Cooling tile pattern: ~36 C commands (9 circles × 4 curves)
    # and ~10 M commands (9 for circles + 1 for border)
    if 30 <= c_count <= 40 and 8 <= m_count <= 12 and z_count >= 1:
        return True
    
    return False

def clone_element_deep(element):
    """Create a deep copy of an element and all its children"""
    # Create a new element with the same tag
    new_elem = ET.Element(element.tag)
    
    # Copy all attributes
    for key, value in element.attrib.items():
        new_elem.set(key, value)
    
    # Copy text content
    if element.text:
        new_elem.text = element.text
    if element.tail:
        new_elem.tail = element.tail
    
    # Recursively copy all children
    for child in element:
        new_child = clone_element_deep(child)
        new_elem.append(new_child)
    
    return new_elem

def is_floor_shape(element):
    """Check if an element is the floor shape based on its characteristics"""
    # Check if it's a group element
    if not element.tag.endswith('g'):
        return False
    
    # Check fill color - floor is silver
    fill = element.get('fill', '')
    if 'silver' not in fill.lower():
        return False
    
    # Check if it has a large path child
    for child in element:
        if child.tag.endswith('path'):
            path_data = child.get('d', '')
            # Floor paths tend to be very long with many coordinates
            if len(path_data) > 200:  # Arbitrary threshold for complex paths
                # Check for large coordinate values typical of floor boundaries
                numbers = re.findall(r'\d+', path_data)
                if numbers:
                    max_num = max(int(n) for n in numbers if n.isdigit())
                    if max_num > 1000:  # Floor typically spans large coordinates
                        return True
    
    return False

def find_floor_shape(root):
    """Find the floor shape in the SVG"""
    for element in root.iter():
        if is_floor_shape(element):
            return element
    return None

def find_all_cooling_tiles(root):
    """Find all individual cooling tiles in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    cooling_tiles = []
    processed_paths = set()
    
    # Strategy: Find all paths that match cooling tile pattern
    for element in root.iter():
        if element.tag.endswith('path'):
            path_data = element.get('d', '')
            path_id = id(element)
            
            if path_id not in processed_paths and is_cooling_tile_pattern(path_data):
                # Check if this is part of a server rack (skip if parent group is red)
                parent = None
                parent_map = {c: p for p in root.iter() for c in p}
                parent = parent_map.get(element)
                
                skip = False
                while parent is not None:
                    if parent.tag.endswith('g'):
                        g_fill = parent.get('fill', '')
                        if 'rgb(230,0,0)' in g_fill:
                            skip = True
                            break
                    parent = parent_map.get(parent)
                
                if not skip:
                    processed_paths.add(path_id)
                    x, y = get_element_position(element, root)
                    
                    # Find the immediate parent group
                    parent_map = {c: p for p in root.iter() for c in p}
                    parent_g = parent_map.get(element)
                    while parent_g is not None and not parent_g.tag.endswith('g'):
                        parent_g = parent_map.get(parent_g)
                    
                    # Get dimensions
                    width, height = get_element_bounds(parent_g if parent_g is not None else element)
                    
                    cooling_tiles.append({
                        'element': parent_g if parent_g is not None else element,
                        'path_element': element,
                        'x': x,
                        'y': y,
                        'width': width,
                        'height': height
                    })
    
    # Remove duplicates based on position (tiles at same location)
    unique_tiles = []
    positions_seen = set()
    
    for tile in cooling_tiles:
        pos_key = f"{tile['x']:.1f},{tile['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_tiles.append(tile)
    
    return unique_tiles

def find_all_cooling_units(root):
    """Find all individual cooling units in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    cooling_units = []
    processed_groups = set()
    
    # Find groups that match cooling unit pattern
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        # Check color - cooling units are blue
        g_fill = g.get('fill', '')
        g_stroke = g.get('stroke', '')
        
        is_blue = ('rgb(62,153,223)' in g_fill or 'rgb(62,153,223)' in g_stroke or
                   'rgb(171,211,241)' in g_fill or 'rgb(171,211,241)' in g_stroke)
        
        if not is_blue:
            continue
        
        # Count direct children that are lines and paths
        has_line = False
        has_path = False
        
        for child in g:
            if child.tag.endswith('line'):
                has_line = True
            elif child.tag.endswith('path'):
                # Check if path has the cooling unit pattern (filled rectangle)
                path_fill = child.get('fill', '')
                if 'rgb(171,211,241)' in path_fill or path_fill != 'none':
                    has_path = True
        
        if has_line and has_path:
            processed_groups.add(g_id)
            
            # Get position
            x, y = get_element_position(g, root)
            
            # Get dimensions
            width, height = get_element_bounds(g)
            
            cooling_units.append({
                'element': g,
                'x': x,
                'y': y,
                'width': width,
                'height': height
            })
            
            # Mark child groups as processed
            for child_g in g.findall('.//svg:g', namespaces):
                processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_units = []
    positions_seen = set()
    
    for unit in cooling_units:
        pos_key = f"{unit['x']:.1f},{unit['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_units.append(unit)
    
    return unique_units

def find_all_server_racks(root):
    """Find all individual server racks in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    server_racks = []
    processed_groups = set()
    
    # Find groups that contain both line and path elements
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        # Check color
        g_fill = g.get('fill', '')
        g_stroke = g.get('stroke', '')
        
        is_red = 'rgb(230,0,0)' in g_fill or 'rgb(230,0,0)' in g_stroke
        is_gray = 'rgb(110,110,110)' in g_fill or 'rgb(110,110,110)' in g_stroke
        
        if not (is_red or is_gray):
            continue
        
        # Count direct children that are lines and paths
        has_line = False
        has_path = False
        
        for child in g:
            if child.tag.endswith('line'):
                has_line = True
            elif child.tag.endswith('path'):
                has_path = True
        
        if has_line and has_path:
            processed_groups.add(g_id)
            
            # Get position
            x, y = get_element_position(g, root)
            
            # Get dimensions
            width, height = get_element_bounds(g)
            
            server_racks.append({
                'element': g,
                'x': x,
                'y': y,
                'color': 'red' if is_red else 'gray',
                'width': width,
                'height': height
            })
            
            # Mark child groups as processed
            for child_g in g.findall('.//svg:g', namespaces):
                processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_racks = []
    positions_seen = set()
    
    for rack in server_racks:
        pos_key = f"{rack['color']}:{rack['x']:.1f},{rack['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_racks.append(rack)
    
    return unique_racks

def find_grid_labels(root):
    """Find all grid label groups (coordinates) in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    grid_groups = []
    
    # Find groups that contain multiple text elements
    for g in root.findall('.//svg:g', namespaces):
        # Count text elements in this group
        text_elements = g.findall('.//svg:text', namespaces)
        
        if len(text_elements) > 10:  # Grid labels typically have many text elements
            # Check if it looks like grid labels (short text content)
            is_grid = True
            for text in text_elements[:5]:  # Check first few
                if text.text and len(text.text) > 4:  # Grid labels are short (e.g., "AA", "01")
                    is_grid = False
                    break
            
            if is_grid:
                grid_groups.append(g)
    
    return grid_groups

def is_wall(element):
    """Check if an element is a wall based on its characteristics"""
    # Check if it's a group element
    if not element.tag.endswith('g'):
        return False
    
    # Check for path children with no fill (walls are typically unfilled paths)
    for child in element:
        if child.tag.endswith('path'):
            fill = child.get('fill', '')
            if fill == 'none':
                # Check if path data contains wall-like patterns (L commands)
                path_data = child.get('d', '')
                if path_data:
                    # Count commands
                    l_count = path_data.upper().count('L')
                    
                    # Walls typically have line segments
                    if l_count >= 1:
                        return True
    
    return False

def is_door(element):
    """Check if an element is a door based on its characteristics"""
    # Check if it's a group element
    if not element.tag.endswith('g'):
        return False
    
    # Check fill color - doors are black with opacity
    g_fill = element.get('fill', '')
    fill_opacity = element.get('fill-opacity', '1')
    
    # Check for black color
    is_black = ('rgb(0,0,0)' in g_fill or 'black' in g_fill.lower())
    
    # Check for semi-transparent fill
    try:
        opacity = float(fill_opacity)
        is_semi_transparent = 0.5 <= opacity <= 0.8
    except:
        is_semi_transparent = False
    
    if not (is_black and is_semi_transparent):
        return False
    
    # Check for path children with arc commands (typical for door swing)
    has_arc = False
    has_rect = False
    
    for child in element:
        if child.tag.endswith('path'):
            path_data = child.get('d', '')
            # Door paths typically contain A (arc) commands for the swing
            if 'A' in path_data.upper():
                has_arc = True
            # And may have rectangular parts
            if 'Z' in path_data and ('L' in path_data or 'M' in path_data):
                has_rect = True
    
    return has_arc  # Door must have an arc

def is_column(element):
    """Check if an element is a column based on its characteristics"""
    # Check if it's a group element
    if not element.tag.endswith('g'):
        return False
    
    # Check fill color - columns are gray
    g_fill = element.get('fill', '')
    g_stroke = element.get('stroke', '')
    
    is_gray = ('rgb(230,230,230)' in g_fill or 'rgb(230,230,230)' in g_stroke or
               'rgb(145,145,145)' in g_fill or 'rgb(145,145,145)' in g_stroke)
    
    if not is_gray:
        return False
    
    # Check for path children that form rectangles
    path_count = 0
    for child in element:
        if child.tag.endswith('path'):
            path_count += 1
            path_data = child.get('d', '')
            # Column paths typically contain Z (close path) and form rectangles
            if 'Z' in path_data:
                return True
    
    # Columns typically have 1-2 paths
    if 1 <= path_count <= 2:
        return True
    
    return False

def is_pdu_or_ppc(element):
    """Check if an element is a PDU or PPC based on its characteristics"""
    # Check if it's a group element
    if not element.tag.endswith('g'):
        return False
    
    # Check fill color - PDUs/PPCs are green
    g_fill = element.get('fill', '')
    g_stroke = element.get('stroke', '')
    
    is_green = ('rgb(103,203,51)' in g_fill or 'rgb(103,203,51)' in g_stroke or
                'rgb(189,232,167)' in g_fill or 'rgb(189,232,167)' in g_stroke)
    
    if not is_green:
        return False
    
    # Check for line and path children (similar structure to cooling units)
    has_line = False
    has_path = False
    
    for child in element:
        if child.tag.endswith('line'):
            has_line = True
        elif child.tag.endswith('path'):
            has_path = True
    
    return has_line and has_path

def find_pdus_ppcs(root):
    """Find all PDUs and PPCs in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    pdus_ppcs = []
    processed_groups = set()
    
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        if is_pdu_or_ppc(g):
            processed_groups.add(g_id)
            
            # Get position
            x, y = get_element_position(g, root)
            
            # Get dimensions
            width, height = get_element_bounds(g)
            
            pdus_ppcs.append({
                'element': g,
                'x': x,
                'y': y,
                'width': width,
                'height': height
            })
            
            # Mark child groups as processed
            for child_g in g.findall('.//svg:g', namespaces):
                processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_pdus_ppcs = []
    positions_seen = set()
    
    for pdu_ppc in pdus_ppcs:
        pos_key = f"{pdu_ppc['x']:.1f},{pdu_ppc['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_pdus_ppcs.append(pdu_ppc)
    
    return unique_pdus_ppcs

def find_walls(root):
    """Find all walls in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    walls = []
    processed_groups = set()
    
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        if is_wall(g):
            # Check if it's not already identified as another type
            g_fill = g.get('fill', '')
            
            # Skip if it's colored like other identified elements
            is_cooling = ('rgb(62,153,223)' in g_fill or 'rgb(171,211,241)' in g_fill)
            is_server = ('rgb(230,0,0)' in g_fill or 'rgb(110,110,110)' in g_fill)
            is_floor = ('silver' in g_fill.lower())
            is_column = ('rgb(230,230,230)' in g_fill or 'rgb(145,145,145)' in g_fill)
            is_pdu = ('rgb(103,203,51)' in g_fill or 'rgb(189,232,167)' in g_fill)
            is_row = ('rgb(224,224,215)' in g_fill)
            is_door = ('rgb(0,0,0)' in g_fill)
            
            if not (is_cooling or is_server or is_floor or is_column or is_pdu or is_row or is_door):
                processed_groups.add(g_id)
                
                # Get position
                x, y = get_element_position(g, root)
                
                # Get dimensions
                width, height = get_element_bounds(g)
                
                walls.append({
                    'element': g,
                    'x': x,
                    'y': y,
                    'width': width,
                    'height': height
                })
                
                # Mark child groups as processed
                for child_g in g.findall('.//svg:g', namespaces):
                    processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_walls = []
    positions_seen = set()
    
    for wall in walls:
        pos_key = f"{wall['x']:.1f},{wall['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_walls.append(wall)
    
    return unique_walls

def find_doors(root):
    """Find all doors in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    doors = []
    processed_groups = set()
    
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        if is_door(g):
            processed_groups.add(g_id)
            
            # Get position
            x, y = get_element_position(g, root)
            
            # Get dimensions
            width, height = get_element_bounds(g)
            
            # Extract rotation if present
            transform = g.get('transform', '')
            rotate_match = re.search(r'rotate\s*\(\s*([+-]?\d*\.?\d+)', transform)
            rotation = float(rotate_match.group(1)) if rotate_match else 0
            
            doors.append({
                'element': g,
                'x': x,
                'y': y,
                'width': width,
                'height': height,
                'rotation': rotation
            })
            
            # Mark child groups as processed
            for child_g in g.findall('.//svg:g', namespaces):
                processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_doors = []
    positions_seen = set()
    
    for door in doors:
        pos_key = f"{door['x']:.1f},{door['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_doors.append(door)
    
    return unique_doors

def find_columns(root):
    """Find all columns in the SVG"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    columns = []
    processed_groups = set()
    
    for g in root.findall('.//svg:g', namespaces):
        g_id = id(g)
        
        if g_id in processed_groups:
            continue
        
        if is_column(g):
            processed_groups.add(g_id)
            
            # Get position
            x, y = get_element_position(g, root)
            
            # Get dimensions
            width, height = get_element_bounds(g)
            
            columns.append({
                'element': g,
                'x': x,
                'y': y,
                'width': width,
                'height': height
            })
            
            # Mark child groups as processed
            for child_g in g.findall('.//svg:g', namespaces):
                processed_groups.add(id(child_g))
    
    # Remove duplicates based on position
    unique_columns = []
    positions_seen = set()
    
    for column in columns:
        pos_key = f"{column['x']:.1f},{column['y']:.1f}"
        if pos_key not in positions_seen:
            positions_seen.add(pos_key)
            unique_columns.append(column)
    
    return unique_columns

def create_patterns_svg(original_root, cooling_tiles, server_racks, cooling_units, floor_shape, grid_labels, walls, columns, pdus_ppcs, rows, doors, output_filename):
    """Create a new SVG file containing the complete patterns (not just dots)"""
    ET.register_namespace('', 'http://www.w3.org/2000/svg')
    
    new_root = ET.Element('{http://www.w3.org/2000/svg}svg')
    
    # Copy attributes from original SVG
    for attr, value in original_root.attrib.items():
        new_root.set(attr, value)
    
    # Add title
    title = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}title')
    title.text = f'Individual Patterns - {len(cooling_tiles)} cooling tiles, {len(server_racks)} server racks, {len(cooling_units)} cooling units, {len(columns)} columns, {len(pdus_ppcs)} PDUs/PPCs, {len(rows)} rows, {len(walls)} walls, {len(doors)} doors'
    
    # Add floor shape first (as background)
    if floor_shape is not None:
        floor_copy = clone_element_deep(floor_shape)
        floor_copy.set('id', 'floor-shape')
        # Make it semi-transparent so other elements are visible
        floor_copy.set('opacity', '0.3')
        new_root.append(floor_copy)
    
    # Add rows (should be behind most elements but above floor)
    if rows:
        rows_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        rows_group.set('id', 'rows')
        
        for i, row in enumerate(rows):
            # Clone the entire row element
            row_copy = clone_element_deep(row['element'])
            row_copy.set('id', f'row-{i}')
            rows_group.append(row_copy)
    
    # Add walls
    if walls:
        walls_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        walls_group.set('id', 'walls')
        
        for i, wall in enumerate(walls):
            # Clone the entire wall element
            wall_copy = clone_element_deep(wall['element'])
            wall_copy.set('id', f'wall-{i}')
            walls_group.append(wall_copy)
    
    # Add doors
    if doors:
        doors_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        doors_group.set('id', 'doors')
        
        for i, door in enumerate(doors):
            # Clone the entire door element
            door_copy = clone_element_deep(door['element'])
            door_copy.set('id', f'door-{i}')
            doors_group.append(door_copy)
    
    # Add columns
    if columns:
        columns_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        columns_group.set('id', 'columns')
        
        for i, column in enumerate(columns):
            # Clone the entire column element
            column_copy = clone_element_deep(column['element'])
            column_copy.set('id', f'column-{i}')
            columns_group.append(column_copy)
    
    # Add grid labels (coordinates)
    if grid_labels:
        grid_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        grid_group.set('id', 'grid-coordinates')
        
        for i, label_group in enumerate(grid_labels):
            # Clone the entire grid label group
            label_copy = clone_element_deep(label_group)
            label_copy.set('id', f'grid-labels-{i}')
            grid_group.append(label_copy)
    
    # Add cooling tiles - copy the full element structure
    if cooling_tiles:
        cooling_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        cooling_group.set('id', 'cooling-tiles')
        
        for i, tile in enumerate(cooling_tiles):
            # Clone the entire cooling tile element
            tile_copy = clone_element_deep(tile['element'])
            
            # Update the ID
            tile_copy.set('id', f'cooling-tile-{i}')
            
            # Add to the group
            cooling_group.append(tile_copy)
    
    # Add server racks - copy the full element structure
    if server_racks:
        server_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        server_group.set('id', 'server-racks')
        
        for i, rack in enumerate(server_racks):
            # Clone the entire server rack element
            rack_copy = clone_element_deep(rack['element'])
            
            # Update the ID
            rack_copy.set('id', f'server-rack-{i}-{rack["color"]}')
            
            # Add to the group
            server_group.append(rack_copy)
    
    # Add cooling units - copy the full element structure
    if cooling_units:
        cooling_units_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        cooling_units_group.set('id', 'cooling-units')
        
        for i, unit in enumerate(cooling_units):
            # Clone the entire cooling unit element
            unit_copy = clone_element_deep(unit['element'])
            
            # Update the ID
            unit_copy.set('id', f'cooling-unit-{i}')
            
            # Add to the group
            cooling_units_group.append(unit_copy)
    
    # Add PDUs/PPCs - copy the full element structure
    if pdus_ppcs:
        pdus_ppcs_group = ET.SubElement(new_root, '{http://www.w3.org/2000/svg}g')
        pdus_ppcs_group.set('id', 'pdus-ppcs')
        
        for i, pdu_ppc in enumerate(pdus_ppcs):
            # Clone the entire PDU/PPC element
            pdu_ppc_copy = clone_element_deep(pdu_ppc['element'])
            
            # Update the ID
            pdu_ppc_copy.set('id', f'pdu-ppc-{i}')
            
            # Add to the group
            pdus_ppcs_group.append(pdu_ppc_copy)
    
    # Write to file
    tree = ET.ElementTree(new_root)
    ET.indent(tree, space="  ")
    tree.write(output_filename, encoding='utf-8', xml_declaration=True)
    
    return output_filename

def create_xlsx_report(cooling_tiles, server_racks, cooling_units, columns, pdus_ppcs, rows, walls, doors, output_filename):
    """Create an Excel file with individual pattern information"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Individual Patterns"
    
    # Headers
    headers = ['object_type', 'object_id', 'coordinates', 'width', 'height', 'rotation']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    row = 2
    
    # Add rows
    for i, row_elem in enumerate(rows):
        ws.cell(row=row, column=1, value="row")
        ws.cell(row=row, column=2, value=f"row-{i}")
        ws.cell(row=row, column=3, value=f"({row_elem['x']:.2f}, {row_elem['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{row_elem['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{row_elem['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Add walls
    for i, wall in enumerate(walls):
        ws.cell(row=row, column=1, value="wall")
        ws.cell(row=row, column=2, value=f"wall-{i}")
        ws.cell(row=row, column=3, value=f"({wall['x']:.2f}, {wall['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{wall['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{wall['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Add doors
    for i, door in enumerate(doors):
        ws.cell(row=row, column=1, value="door")
        ws.cell(row=row, column=2, value=f"door-{i}")
        ws.cell(row=row, column=3, value=f"({door['x']:.2f}, {door['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{door['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{door['height']:.2f}")
        ws.cell(row=row, column=6, value=f"{door['rotation']:.1f}°")
        row += 1
    
    # Add cooling tiles
    for i, tile in enumerate(cooling_tiles):
        ws.cell(row=row, column=1, value="cooling_tile")
        ws.cell(row=row, column=2, value=f"cooling-tile-{i}")
        ws.cell(row=row, column=3, value=f"({tile['x']:.2f}, {tile['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{tile['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{tile['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Add server racks
    for i, rack in enumerate(server_racks):
        ws.cell(row=row, column=1, value=f"server_rack_{rack['color']}")
        ws.cell(row=row, column=2, value=f"server-rack-{i}")
        ws.cell(row=row, column=3, value=f"({rack['x']:.2f}, {rack['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{rack['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{rack['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Add cooling units
    for i, unit in enumerate(cooling_units):
        ws.cell(row=row, column=1, value="cooling_unit")
        ws.cell(row=row, column=2, value=f"cooling-unit-{i}")
        ws.cell(row=row, column=3, value=f"({unit['x']:.2f}, {unit['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{unit['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{unit['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Add columns
    for i, column in enumerate(columns):
        ws.cell(row=row, column=1, value="column")
        ws.cell(row=row, column=2, value=f"column-{i}")
        ws.cell(row=row, column=3, value=f"({column['x']:.2f}, {column['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{column['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{column['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Add PDUs/PPCs
    for i, pdu_ppc in enumerate(pdus_ppcs):
        ws.cell(row=row, column=1, value="pdu_ppc")
        ws.cell(row=row, column=2, value=f"pdu-ppc-{i}")
        ws.cell(row=row, column=3, value=f"({pdu_ppc['x']:.2f}, {pdu_ppc['y']:.2f})")
        ws.cell(row=row, column=4, value=f"{pdu_ppc['width']:.2f}")
        ws.cell(row=row, column=5, value=f"{pdu_ppc['height']:.2f}")
        ws.cell(row=row, column=6, value="N/A")
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Summary
    row += 2
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total Rows:")
    ws.cell(row=row, column=2, value=len(rows))
    row += 1
    ws.cell(row=row, column=1, value="Total Walls:")
    ws.cell(row=row, column=2, value=len(walls))
    row += 1
    ws.cell(row=row, column=1, value="Total Doors:")
    ws.cell(row=row, column=2, value=len(doors))
    row += 1
    ws.cell(row=row, column=1, value="Total Cooling Tiles:")
    ws.cell(row=row, column=2, value=len(cooling_tiles))
    row += 1
    ws.cell(row=row, column=1, value="Total Server Racks:")
    ws.cell(row=row, column=2, value=len(server_racks))
    row += 1
    ws.cell(row=row, column=1, value="Total Cooling Units:")
    ws.cell(row=row, column=2, value=len(cooling_units))
    row += 1
    ws.cell(row=row, column=1, value="Total Columns:")
    ws.cell(row=row, column=2, value=len(columns))
    row += 1
    ws.cell(row=row, column=1, value="Total PDUs/PPCs:")
    ws.cell(row=row, column=2, value=len(pdus_ppcs))
    
    # Count by color
    red_racks = sum(1 for r in server_racks if r['color'] == 'red')
    gray_racks = sum(1 for r in server_racks if r['color'] == 'gray')
    
    if red_racks > 0:
        row += 1
        ws.cell(row=row, column=1, value="  - Red Server Racks:")
        ws.cell(row=row, column=2, value=red_racks)
    if gray_racks > 0:
        row += 1
        ws.cell(row=row, column=1, value="  - Gray Server Racks:")
        ws.cell(row=row, column=2, value=gray_racks)
    
    # Add coordinate statistics
    row += 2
    ws.cell(row=row, column=1, value="Coordinate Range:").font = Font(bold=True)
    
    if walls:
        x_coords = [w['x'] for w in walls]
        y_coords = [w['y'] for w in walls]
        widths = [w['width'] for w in walls]
        heights = [w['height'] for w in walls]
        row += 1
        ws.cell(row=row, column=1, value="Walls X range:")
        ws.cell(row=row, column=2, value=f"{min(x_coords):.2f} to {max(x_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Walls Y range:")
        ws.cell(row=row, column=2, value=f"{min(y_coords):.2f} to {max(y_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Walls avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if doors:
        x_coords = [d['x'] for d in doors]
        y_coords = [d['y'] for d in doors]
        widths = [d['width'] for d in doors]
        heights = [d['height'] for d in doors]
        row += 1
        ws.cell(row=row, column=1, value="Doors X range:")
        ws.cell(row=row, column=2, value=f"{min(x_coords):.2f} to {max(x_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Doors Y range:")
        ws.cell(row=row, column=2, value=f"{min(y_coords):.2f} to {max(y_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Doors avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if rows:
        x_coords = [r['x'] for r in rows]
        y_coords = [r['y'] for r in rows]
        widths = [r['width'] for r in rows]
        heights = [r['height'] for r in rows]
        row += 1
        ws.cell(row=row, column=1, value="Rows X range:")
        ws.cell(row=row, column=2, value=f"{min(x_coords):.2f} to {max(x_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Rows Y range:")
        ws.cell(row=row, column=2, value=f"{min(y_coords):.2f} to {max(y_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Rows avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if cooling_tiles:
        x_coords = [t['x'] for t in cooling_tiles]
        y_coords = [t['y'] for t in cooling_tiles]
        widths = [t['width'] for t in cooling_tiles]
        heights = [t['height'] for t in cooling_tiles]
        row += 1
        ws.cell(row=row, column=1, value="Cooling Tiles X range:")
        ws.cell(row=row, column=2, value=f"{min(x_coords):.2f} to {max(x_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Cooling Tiles Y range:")
        ws.cell(row=row, column=2, value=f"{min(y_coords):.2f} to {max(y_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Cooling Tiles avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if cooling_units:
        x_coords = [u['x'] for u in cooling_units]
        y_coords = [u['y'] for u in cooling_units]
        widths = [u['width'] for u in cooling_units]
        heights = [u['height'] for u in cooling_units]
        row += 1
        ws.cell(row=row, column=1, value="Cooling Units X range:")
        ws.cell(row=row, column=2, value=f"{min(x_coords):.2f} to {max(x_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Cooling Units Y range:")
        ws.cell(row=row, column=2, value=f"{min(y_coords):.2f} to {max(y_coords):.2f}")
        row += 1
        ws.cell(row=row, column=1, value="Cooling Units avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if server_racks:
        widths = [r['width'] for r in server_racks]
        heights = [r['height'] for r in server_racks]
        row += 1
        ws.cell(row=row, column=1, value="Server Racks avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if columns:
        widths = [c['width'] for c in columns]
        heights = [c['height'] for c in columns]
        row += 1
        ws.cell(row=row, column=1, value="Columns avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    if pdus_ppcs:
        widths = [p['width'] for p in pdus_ppcs]
        heights = [p['height'] for p in pdus_ppcs]
        row += 1
        ws.cell(row=row, column=1, value="PDUs/PPCs avg size:")
        ws.cell(row=row, column=2, value=f"{sum(widths)/len(widths):.2f} x {sum(heights)/len(heights):.2f}")
    
    # Save
    wb.save(output_filename)
    return output_filename

def debug_svg_structure(root):
    """Debug function to understand SVG structure"""
    namespaces = {'svg': 'http://www.w3.org/2000/svg'}
    
    print("\n=== SVG Structure Analysis ===")
    
    # Find some cooling tile patterns and analyze their structure
    count = 0
    for element in root.iter():
        if element.tag.endswith('path'):
            path_data = element.get('d', '')
            if is_cooling_tile_pattern(path_data):
                count += 1
                if count <= 3:  # Analyze first 3
                    print(f"\nCooling Tile Pattern {count}:")
                    
                    # Get parent hierarchy
                    parent_map = {c: p for p in root.iter() for c in p}
                    current = element
                    hierarchy = []
                    
                    while current is not None:
                        tag = current.tag.split('}')[-1] if '}' in current.tag else current.tag
                        transform = current.get('transform', 'none')
                        hierarchy.append(f"{tag} (transform: {transform})")
                        current = parent_map.get(current)
                    
                    print("  Hierarchy:", " -> ".join(reversed(hierarchy)))
                    
                    # Get position
                    x, y = get_element_position(element, root)
                    print(f"  Calculated position: ({x:.2f}, {y:.2f})")

def main():
    # Input SVG filename
    svg_filename = "MLQ2 G.F - Power.svg"
    
    # Parse SVG
    root = parse_svg_file(svg_filename)
    if root is None:
        return
    
    print("\n=== Detecting Individual Patterns ===")
    
    # Debug SVG structure
    debug_svg_structure(root)
    
    # Find patterns
    cooling_tiles = find_all_cooling_tiles(root)
    server_racks = find_all_server_racks(root)
    cooling_units = find_all_cooling_units(root)
    floor_shape = find_floor_shape(root)
    grid_labels = find_grid_labels(root)
    walls = find_walls(root)
    doors = find_doors(root)
    columns = find_columns(root)
    pdus_ppcs = find_pdus_ppcs(root)
    rows = find_rows(root)
    
    print(f"\n=== Results ===")
    print(f"Found {len(rows)} rows")
    print(f"Found {len(walls)} walls")
    print(f"Found {len(doors)} doors")
    print(f"Found {len(cooling_tiles)} individual cooling tiles")
    print(f"Found {len(server_racks)} individual server racks")
    print(f"Found {len(cooling_units)} individual cooling units")
    print(f"Found {len(columns)} columns")
    print(f"Found {len(pdus_ppcs)} PDUs/PPCs")
    print(f"Found floor shape: {'Yes' if floor_shape is not None else 'No'}")
    print(f"Found {len(grid_labels)} grid label groups")
    
    # Show position distribution
    if walls:
        x_coords = [w['x'] for w in walls]
        y_coords = [w['y'] for w in walls]
        print(f"\nWall positions:")
        print(f"  X range: {min(x_coords):.2f} to {max(x_coords):.2f}")
        print(f"  Y range: {min(y_coords):.2f} to {max(y_coords):.2f}")
    
    if doors:
        print(f"\nDoor details:")
        for i, door in enumerate(doors[:5]):  # Show first 5
            print(f"  Door {i}: pos=({door['x']:.2f}, {door['y']:.2f}), rotation={door['rotation']:.1f}°")
    
    if rows:
        x_coords = [r['x'] for r in rows]
        y_coords = [r['y'] for r in rows]
        print(f"\nRow positions:")
        print(f"  X range: {min(x_coords):.2f} to {max(x_coords):.2f}")
        print(f"  Y range: {min(y_coords):.2f} to {max(y_coords):.2f}")
        
        # Show first few unique positions
        unique_positions = set()
        for row in rows[:10]:
            pos = f"({row['x']:.2f}, {row['y']:.2f})"
            unique_positions.add(pos)
        print(f"  First few positions: {', '.join(list(unique_positions)[:5])}")
    
    if cooling_tiles:
        x_coords = [t['x'] for t in cooling_tiles]
        y_coords = [t['y'] for t in cooling_tiles]
        print(f"\nCooling tile positions:")
        print(f"  X range: {min(x_coords):.2f} to {max(x_coords):.2f}")
        print(f"  Y range: {min(y_coords):.2f} to {max(y_coords):.2f}")
        
        # Show first few unique positions
        unique_positions = set()
        for tile in cooling_tiles[:10]:
            pos = f"({tile['x']:.2f}, {tile['y']:.2f})"
            unique_positions.add(pos)
        print(f"  First few positions: {', '.join(list(unique_positions)[:5])}")
    
    if cooling_units:
        x_coords = [u['x'] for u in cooling_units]
        y_coords = [u['y'] for u in cooling_units]
        print(f"\nCooling unit positions:")
        print(f"  X range: {min(x_coords):.2f} to {max(x_coords):.2f}")
        print(f"  Y range: {min(y_coords):.2f} to {max(y_coords):.2f}")
    
    # Create outputs
    output_svg = svg_filename.replace('.svg', '_individual_patterns.svg')
    create_patterns_svg(root, cooling_tiles, server_racks, cooling_units, floor_shape, grid_labels, walls, columns, pdus_ppcs, rows, doors, output_svg)
    print(f"\nCreated SVG with individual patterns: {output_svg}")
    
    output_xlsx = svg_filename.replace('.svg', '_individual_patterns.xlsx')
    create_xlsx_report(cooling_tiles, server_racks, cooling_units, columns, pdus_ppcs, rows, walls, doors, output_xlsx)
    print(f"Created XLSX report: {output_xlsx}")

if __name__ == "__main__":
    main()